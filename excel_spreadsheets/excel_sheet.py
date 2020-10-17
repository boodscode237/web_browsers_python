import openpyxl
import os

os.chdir('/home/abodo/Documents/webbrowser_python/excel_spreadsheets')

workbook = openpyxl.load_workbook('example.xlsx')
sheet = workbook.get_sheet_by_name('Sheet1')
workbook.get_sheet_names()
cell = sheet['C1']
cell.value
print(sheet.cell(row=1, column=2))

# function to read rows
for row in range(2, sheet.max_row + 1):
    # Each row in the spreadsheet has data for one census tract.
    state = sheet['B' + str(row)].value
    county = sheet['C' + str(row)].value
    pop = sheet['D' + str(row)].value


# modify sheets

wb = openpyxl.Workbook()
wb.get_sheet_names()
sheet = wb.get_sheet_by_name('Sheet')
sheet['A1'] = 16547
sheet['B1'] = 'Hello'

wb.save('example1.xlsx')
sheet2 = wb.create_sheet()
sheet2.title = "cool name"
wb.save('example2.xlsx')
